# -*- coding: utf-8 -*-
"""
MODULE_NAME: core/schema.py
PURPOSE: 추출 '데이터 계약' — FieldSchema/Schema. 난독화 클래스와 무관한 '구조 위치(경로/시그니처)'를
         담고, JSON 캐시 및 CSV/엑셀 '레시피'(사람이 열어봐도 읽히는 타격 위치표)로 직렬화/역직렬화한다.
DEPENDENCY: json/csv/dataclasses(표준). 엑셀 레시피에 한해 openpyxl(선택; 지연 import).

[검증된 주요 사이트 및 케이스]
- CSV 레시피 왕복(save_csv_recipe ↔ from_csv_recipe): 사이트별 자동 저장/재현의 핵심(replay 번호 재현).
- 체인 크롤링: extra_meta(url_col/clean_url/chain 등)를 meta 행으로 함께 저장 → 입력 없이 재현.
- 단일 레코드(single_record): 상세페이지를 1행으로 추출(체인 상세 스키마).

[테스트/운영 교훈]
- 이 계층은 순수 데이터/직렬화다(DOM/네트워크/LLM 없음) → engine 을 import 하지 않는다(leaf).
- meta 는 CSV 의 tag 열에 값을 담는 규약(kind=meta 행) — 사람이 열어도 읽히도록 한 파일에 통합.
- [이력] Phase 3 에서 engine.py 에서 이관(동작 0 변경). engine 은 core.schema 를 import 해 재사용.
"""
from __future__ import annotations

import json
from dataclasses import dataclass, field, asdict
from typing import Optional


@dataclass
class FieldSchema:
    css: str                 # 빠른 경로용 클래스 기반 셀렉터 (예: a.post-title)
    tag: str
    cls: Optional[str]
    path: list               # 구조 경로(치유 anchor) [(tag, idx), ...]
    attr: Optional[str]
    example: str = ""        # 캘리브레이션 시점의 예시 값 (LLM 치유의 참조용)
    # (v5.0) 한 노드에 여러 필드가 뭉쳤을 때 '사용자 예시 경계'로 텍스트를 쪼갠다(사이트 무관).
    # seg_seps = 세그먼트 사이 구분자 리스트(예시에서 파생), seg_index = 이 필드의 조각 번호.
    seg_index: Optional[int] = None
    seg_seps: Optional[list] = None


@dataclass
class Schema:
    row_css: str
    row_tag: str
    row_cls: Optional[str]
    row_signature: str
    # 반복 레코드의 '컴포넌트 경계 속성'(예: 'data-testid=trend', 'role=listitem').
    # 클래스가 균일한 SPA(X 등)에서 구조 시그니처가 과다매칭될 때, 이 속성으로
    # 정확히 레코드만 추린다(있으면 최우선 가속 경로). 형식: "<attr>=<value>".
    row_testid: Optional[str] = None
    # 단일 레코드 모드: 상세 페이지처럼 '반복 목록'이 아니라 페이지당 레코드 1개인 경우.
    # 체인 크롤링(부모 CSV 링크 → 상세 페이지)에서 각 상세 페이지를 1행으로 추출할 때 켠다.
    # 켜지면 extract 가 MIN_ROWS(반복 2개) 요구를 건너뛰고 컨테이너 1개만 잡는다.
    single_record: bool = False
    fields: dict = field(default_factory=dict)  # name -> FieldSchema(dict)

    def save(self, path: str):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(asdict(self), f, ensure_ascii=False, indent=2)

    @staticmethod
    def load(path: str) -> "Schema":
        with open(path, "r", encoding="utf-8") as f:
            d = json.load(f)
        s = Schema(d["row_css"], d["row_tag"], d["row_cls"],
                   d["row_signature"], d.get("row_testid"))
        s.single_record = bool(d.get("single_record", False))
        s.fields = d["fields"]
        return s

    # --- 엑셀 레시피(사람이 보고 참조 가능한 '엘레멘트 주소록') ---------------
    def save_excel(self, path: str, url: str = ""):
        """스키마(필드별 위치/경로)를 엑셀로 저장. 값이 아니라 '참조 대상'을 기록."""
        import openpyxl
        wb = openpyxl.Workbook()
        m = wb.active
        m.title = "meta"
        m.append(["key", "value"])
        for k, v in [("url", url), ("row_tag", self.row_tag),
                     ("row_cls", self.row_cls or ""), ("row_css", self.row_css),
                     ("row_signature", self.row_signature)]:
            m.append([k, v])
        f = wb.create_sheet("fields")
        f.append(["name", "tag", "cls", "attr", "path", "example(힌트)"])
        for name, fs in self.fields.items():
            f.append([name, fs["tag"], fs["cls"] or "", fs["attr"] or "",
                      json.dumps(fs["path"], ensure_ascii=False),
                      fs.get("example", "")])
        wb.save(path)

    # --- CSV 레시피(사이트별 자동 저장/재현용 '타격 위치' 기록) ----------------
    def save_csv_recipe(self, path: str, url: str = "", load_method: str = "auto",
                        wait: int = 0, pages: int = 1, extra_meta: dict = None):
        """스키마를 CSV 한 파일로 저장. 값이 아니라 '랜덤클래스 무관 구조 위치'를 기록.

        한 파일에 meta 행과 field 행을 kind 열로 구분해 담는다(사람이 열어봐도 읽힘).
        columns: kind, name, tag, cls, attr, path(json), example
          - kind=meta : name=<키>, tag=<값>  (url/row_*/load_method)
          - kind=field: 필드별 위치(tag/cls/attr/구조경로/예시힌트)
        load_method='chrome' 이면 재현 시 곧장 '진짜 크롬'으로 로드한다(쿠팡 등).
        extra_meta: 체인 크롤링 재현에 필요한 추가 메타(url_col/clean_url/chain 등).
        """
        import csv
        import safe_io
        # 엑셀 등이 레시피 CSV 를 열어 잠갔으면 풀릴 때까지 대기(구멍 방지).
        with safe_io.open_when_writable(path, "w", encoding="utf-8-sig", newline="") as fp:
            w = csv.writer(fp)
            w.writerow(["kind", "name", "tag", "cls", "attr", "path", "example",
                        "seg_index", "seg_seps"])
            meta = [("url", url), ("row_tag", self.row_tag),
                    ("row_cls", self.row_cls or ""), ("row_css", self.row_css),
                    ("row_signature", self.row_signature),
                    ("row_testid", self.row_testid or ""),
                    ("single_record", "1" if self.single_record else "0"),
                    ("load_method", load_method or "auto"),
                    ("wait", str(int(wait or 0))),
                    ("pages", str(int(pages or 1)))]
            for k, v in list(extra_meta.items()) if extra_meta else []:
                meta.append((k, "" if v is None else str(v)))
            for k, v in meta:
                w.writerow(["meta", k, v, "", "", "", "", "", ""])
            for name, fs in self.fields.items():
                si = fs.get("seg_index")
                ss = fs.get("seg_seps")
                w.writerow(["field", name, fs["tag"], fs["cls"] or "",
                            fs["attr"] or "",
                            json.dumps(fs["path"], ensure_ascii=False),
                            fs.get("example", ""),
                            "" if si is None else str(si),
                            json.dumps(ss, ensure_ascii=False) if ss else ""])

    @staticmethod
    def read_recipe_meta(path: str) -> dict:
        """레시피의 meta 행만 dict 로 읽는다(url_col/clean_url/chain 등 조회용)."""
        import csv
        meta = {}
        with open(path, "r", encoding="utf-8-sig", newline="") as fp:
            for row in csv.reader(fp):
                if not row or row[0] != "meta":
                    continue
                cells = (row + [""] * 7)[:7]
                meta[cells[1]] = cells[2]
        return meta

    @staticmethod
    def from_csv_recipe(path: str):
        """CSV 레시피를 읽어 (Schema, url, load_method, wait, pages) 반환."""
        import csv
        meta, fields = {}, {}
        with open(path, "r", encoding="utf-8-sig", newline="") as fp:
            for row in csv.reader(fp):
                if not row or row[0] == "kind":
                    continue
                kind = row[0]
                cells = (row + [""] * 9)[:9]
                _, name, tag, cls, attr, pth, example, seg_i, seg_s = cells
                if kind == "meta":
                    meta[name] = tag    # meta 는 tag 열에 값을 담아둠
                elif kind == "field" and name:
                    fields[name] = asdict(FieldSchema(
                        css=(f"{tag}.{cls}" if cls else tag), tag=tag,
                        cls=cls or None,
                        path=json.loads(pth) if pth else [],
                        attr=attr or None, example=example or "",
                        seg_index=(int(seg_i) if str(seg_i).strip() not in ("", "None")
                                   else None),
                        seg_seps=(json.loads(seg_s) if str(seg_s).strip() else None)))
        s = Schema(meta.get("row_css", ""), meta.get("row_tag", ""),
                   meta.get("row_cls") or None, meta.get("row_signature", ""),
                   meta.get("row_testid") or None)
        s.single_record = str(meta.get("single_record", "0")).strip() in ("1", "true", "True")
        s.fields = fields
        try:
            _wait = int(meta.get("wait", "0") or 0)
        except ValueError:
            _wait = 0
        try:
            _pages = int(meta.get("pages", "1") or 1)
        except ValueError:
            _pages = 1
        return (s, meta.get("url", ""), meta.get("load_method", "auto"),
                _wait, _pages)

    @staticmethod
    def from_excel(path: str):
        """엑셀 레시피를 읽어 (Schema, url) 반환."""
        import openpyxl
        wb = openpyxl.load_workbook(path)
        meta = {r[0]: (r[1] if r[1] is not None else "")
                for r in wb["meta"].iter_rows(min_row=2, values_only=True) if r and r[0]}
        s = Schema(meta.get("row_css", ""), meta.get("row_tag", ""),
                   meta.get("row_cls") or None, meta.get("row_signature", ""))
        for r in wb["fields"].iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            name, tag, cls, attr, path, example = (list(r) + [None] * 6)[:6]
            s.fields[name] = asdict(FieldSchema(
                css=(f"{tag}.{cls}" if cls else tag), tag=tag, cls=cls or None,
                path=json.loads(path) if path else [], attr=attr or None,
                example=example or ""))
        return s, meta.get("url", "")
